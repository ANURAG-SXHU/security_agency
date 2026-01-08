import frappe
from frappe.utils import getdate, add_days
from datetime import date
import calendar


# =========================================================
# MAIN DATA SOURCE (USED BY TABLE + CALENDAR + EXCEL)
# =========================================================
@frappe.whitelist()
def get_shift_calendar(site, month):
    """
    Generate planned shift calendar
    Rotation is carried forward week-by-week across months
    SAFE DEBUG VERSION (Frappe v15 compatible)
    """

    # ---------------- GUARDS ----------------
    if not site:
        frappe.throw("Site is required")

    if not month:
        frappe.throw("Month is required")

    month_date = getdate(month)
    year = month_date.year
    month_num = month_date.month

    month_start = date(year, month_num, 1)
    month_end = date(year, month_num, calendar.monthrange(year, month_num)[1])

    # ---------------- SAFE DEBUG ----------------
    frappe.log_error(
        f"Site={site}, Month={month}, Start={month_start}, End={month_end}",
        "SHIFT CALENDAR DEBUG: Month Info"
    )

    rotations = frappe.get_all(
        "Guard Shift Rotation",
        filters={"site": site},
        fields=[
            "name",
            "guard",
            "day_of_week",
            "rotation_start_date"
        ]
    )

    frappe.log_error(
        f"Rotations found={len(rotations)}",
        "SHIFT CALENDAR DEBUG: Rotations Found"
    )

    results = []

    for rot in rotations:
        rot_doc = frappe.get_doc("Guard Shift Rotation", rot.name)

        if not rot_doc.guard_shift_rotation_item:
            frappe.log_error(
                f"Rotation {rot.name} has no rotation items",
                "SHIFT CALENDAR DEBUG: Empty Rotation"
            )
            continue

        sequence = sorted(
            rot_doc.guard_shift_rotation_item,
            key=lambda r: r.order
        )

        # ---------------- ROTATION LOG ----------------
        frappe.log_error(
            f"Rotation={rot.name}, Guard={rot.guard}, "
            f"Weekday={rot.day_of_week}, "
            f"Start={rot.rotation_start_date}, "
            f"Sequence={len(sequence)}",
            "SHIFT CALENDAR DEBUG: Rotation Details"
        )

        # 🔑 Start from rotation start date
        d = rot.rotation_start_date

        # Align weekday
        while d.strftime("%A") != rot.day_of_week:
            d = add_days(d, 1)

        index = 0

        while d <= month_end:

            if d >= month_start:
                shift = sequence[index % len(sequence)].shift_type

                guard_name = frappe.db.get_value(
                    "Employee",
                    rot.guard,
                    "employee_name"
                )

                results.append({
                    "guard": guard_name,
                    "date": str(d),
                    "shift": shift
                })

            index += 1
            d = add_days(d, 7)

    # ---------------- FINAL DEBUG ----------------
    frappe.log_error(
        f"Total shifts generated={len(results)}",
        "SHIFT CALENDAR DEBUG: Final Results"
    )

    return results
@frappe.whitelist()
def export_shift_calendar_excel(site, month):
    """
    Export Shift Calendar to Excel (Calendar Layout)
    SAFE VERSION – handles None shifts
    """

    if not site:
        frappe.throw("Site is required")

    if not month:
        frappe.throw("Month is required")

    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from frappe.utils import getdate
    from datetime import date
    import calendar
    from io import BytesIO

    # ---------------- GET DATA ----------------
    data = get_shift_calendar(site, month)
    if not data:
        frappe.throw("No shift data found for selected month")

    month_date = getdate(month)
    year = month_date.year
    month_num = month_date.month

    # ---------------- WORKBOOK ----------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Shift Calendar"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---------------- STYLES ----------------
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)

    shift_fills = {
        "A SHIFT": PatternFill("solid", fgColor="C6EFCE"),  # green
        "B SHIFT": PatternFill("solid", fgColor="BDD7EE"),  # blue
        "C SHIFT": PatternFill("solid", fgColor="E4D7F5"),  # purple
        "UNASSIGNED": PatternFill("solid", fgColor="F8CBAD")  # red (optional)
    }

    # ---------------- TITLE ----------------
    title = f"Shift Calendar - {site} ({month_date.strftime('%B %Y')})"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    # ---------------- HEADERS ----------------
    days = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    for col, day in enumerate(days, start=1):
        cell = ws.cell(row=3, column=col, value=day)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[chr(64 + col)].width = 30

    # ---------------- GROUP DATA BY DATE ----------------
    data_map = {}
    for r in data:
        date_key = r.get("date")
        if date_key:
            data_map.setdefault(date_key, []).append(r)

    cal = calendar.monthcalendar(year, month_num)
    start_row = 4

    for week in cal:
        for col, day in enumerate(week, start=1):
            cell = ws.cell(row=start_row, column=col)
            cell.alignment = center

            if day == 0:
                continue

            date_str = f"{year}-{month_num:02d}-{day:02d}"
            lines = [str(day)]

            if date_str in data_map:
                for r in data_map[date_str]:
                    guard = r.get("guard") or "Unknown Guard"
                    shift_value = r.get("shift") or "UNASSIGNED"

                    lines.append(f"{guard} ({shift_value})")

                    for shift_key, fill in shift_fills.items():
                        if shift_key in shift_value:
                            cell.fill = fill

            cell.value = "\n".join(lines)

        start_row += 1

    # ---------------- RESPONSE ----------------
    file_stream = BytesIO()
    wb.save(file_stream)

    frappe.response["filename"] = f"Shift_Calendar_{site}_{month}.xlsx"
    frappe.response["filecontent"] = file_stream.getvalue()
    frappe.response["type"] = "binary"
